from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd


def ATMlog_data():

    # 读取埋点源 excel 文件
    source_excel_file = "I:/Buried-point-testing/ATMlog.xlsx"

    source_workbook = load_workbook(excel_file)

    source_sheet = source_workbook.worksheets[0]

    # 加载目标 excel 文件
    excel_file2 = 'I:/Buried-point-testing/jk埋点.xlsx'

    target_workbook = load_workbook(excel_file2)

    # 创建一个名为“埋点数据”的新工作表

    new_sheet = target_workbook.create_sheet(title="埋点数据")

    target_sheet = target_workbook.sheetnames

    # 选择埋点数据表
    summary_table_sheet = target_workbook[target_sheet[2]]

    # 从源表中取出数据并放置到目标表中
    for row in source_sheet.iter_rows():
        row_data = [cell.value for cell in row]
        summary_table_sheet.append(row_data)

    # 删除第三列
    summary_table_sheet.delete_cols(3)

    # 提取第五列数据
    # fiveth_column = [cell.value for col in summary_table_sheet.iter_cols(
    #     min_col=5, max_col=5) for cell in col]

    # 保存工作簿，包含整理后的数据
    output_file = 'I:/Buried-point-testing/jk埋点.xlsx'
    target_workbook.save(output_file)


# ATMlog_data()


def result_table(fiveth_column):
    '''处理查询出来的埋点数据，整理成结果表'''
    # 加载目标 excel 文件
    excel_file2 = 'I:/Buried-point-testing/jk埋点.xlsx'

    target_workbook = load_workbook(excel_file2)

    # target_sheet = target_workbook.active

    # 创建一个名为“QQ埋点数据”的新工作表
    new_sheet = target_workbook.create_sheet(title="QQ埋点结果表")

    sheets = target_workbook.sheetnames

    # 选择源数据表和目标数据表
    source_sheet = target_workbook.worksheets[0]
    target_sheet = target_workbook.worksheets[3]

    # 提取第一个工作表的第1、2、4列数据
    column1_data = [cell.value for cell in source_sheet['A']]
    column2_data = [cell.value for cell in source_sheet['B']]
    column4_data = [cell.value for cell in source_sheet['D']]

    # 将数据放置到第三个工作表的第1、2、3列中
    for i, value in enumerate(column1_data):
        target_sheet.cell(row=i+1, column=1, value=value)

    for i, value in enumerate(column2_data):
        target_sheet.cell(row=i+1, column=2, value=value)

    for i, value in enumerate(column4_data):
        target_sheet.cell(row=i+1, column=3, value=value)

    # 保存工作簿，包含整理后的数据
    output_file = 'I:/Buried-point-testing/jk埋点.xlsx'
    target_workbook.save(output_file)

# result_table()


def find_data():
    '''查找埋点数据中的需要的数据'''
    # 读取JK埋点 excel 文件
    excel_file2 = '/Users/xinwang/Desktop/zhengsiyu/Buried-point-testing/jk埋点.xlsx'

    target_workbook = load_workbook(excel_file2)

    # 创建一个名为“QQ埋点数据”的新工作表
    new_sheet = target_workbook.create_sheet(title="临时中转表")

    # 选取埋点数据工作表
    data_sheet = target_workbook.worksheets[2]

    # 选取要匹配的数据
    match_data_sheet = target_workbook.worksheets[3]

    # 用于存储第二列的值
    second_column_values = []

    # 遍历第二列的单元格
    for cell in match_data_sheet.iter_cols(min_col=2, max_col=2, values_only=True):
        second_column_values = list(filter(None, cell))[1:]

    # 创建一个空列表用于存储再列表中
    title_list = []

    data_list = []

    for i, row in enumerate(data_sheet.iter_rows(values_only=True)):
        if i == 0:
            title_list = list(row)
        else:
            data_list.append(list(row))

    matched_list = []

    for source_item in second_column_values:
        for target_row in data_list:
            if source_item in target_row:
                matched_list.append(target_row)
                break

    return '全部数据'


find_data()


def second_data:


first_data = find_data()

second_data = second_data(first_data)


third_data = third_data(second_data)
