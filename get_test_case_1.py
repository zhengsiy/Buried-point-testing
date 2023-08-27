from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from get_jk_path import get_jk_path
import pandas as pd


def get_traspose_data():
    '''编写埋点用例'''

    # 读取 excel 文件
    excel_file = get_jk_path()

    wb = load_workbook(excel_file)

    # 创建一个名为“埋点用例”的新工作表
    new_sheet = wb.create_sheet(title='埋点用例')

    # 获取所有工作表的名称
    sheet_names = wb.sheetnames

    # 选择第1个工作表
    selected_sheet = wb[sheet_names[0]]

    # 获取数据的起始行和结束行
    start_row = selected_sheet.min_row
    end_row = selected_sheet.max_row

    # 获取数据的起始列和结束列
    start_column = selected_sheet.min_column
    end_column = selected_sheet.max_column

    # 选择指定范围内的数据（从1开始）
    # minx_row：起始行数
    # max_row：结束行数
    # min_col：起始列数
    # max_col：结束列数

    # 获取第一行的数据
    first_row_values = [
        cell.value for cell in selected_sheet[start_row] if cell.value]

    selected_data = []

    for row in selected_sheet.iter_rows(min_row=start_row+1, max_row=end_row, min_col=start_column,
                                        max_col=end_column, values_only=True):

        selected_data.append(first_row_values)

        selected_data.append(row)

    # 将数据转换为DataFrame并进行转置
    df = pd.DataFrame(selected_data)
    transposed_df = df.transpose()

    # 将转置后的数据粘贴到新的工作表中
    for row in dataframe_to_rows(transposed_df, index=False, header=False):
        new_sheet.append(row)

    # 保存工作簿
    output_file = "I:/Buried-point-testing/jk埋点.xlsx"  # 替换为您的输出Excel文件路径
    wb.save(output_file)

    return


get_traspose_data()
