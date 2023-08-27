from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd
from get_jk_path import get_jk_path


def result_traspose():  # 将最后的结果表进行转置
    '''将最后的结果表进行转置'''
    # 读取 excel 文件
    excel_file = get_jk_path()

    wb = load_workbook(excel_file)

    # 选择第1个工作表
    selected_sheet = wb.worksheets[3]

    # 获取数据的起始行和结束行
    start_row = selected_sheet.min_row
    end_row = selected_sheet.max_row

    # 获取数据的起始列和结束列
    start_column = selected_sheet.min_column
    end_column = selected_sheet.max_column

    # 获取第一行的数据
    first_row_values = [
        cell.value for cell in selected_sheet[start_row] if cell.value]

    selected_data = []

    for row in selected_sheet.iter_rows(min_row=start_row+1, max_row=end_row, min_col=start_column,
                                        max_col=end_column, values_only=True):
        selected_data.append(list(first_row_values))
        selected_data.append(list(row))  # 追加当前行的数据

    # 将数据转换为DataFrame并进行转置
    df = pd.DataFrame(selected_data)
    transposed_df = df.transpose()

    # 将转置后的数据插入到指定行
    target_row = end_row + 5
    for i, row_data in enumerate(dataframe_to_rows(transposed_df, index=False, header=False), start=1):
        for col_idx, value in enumerate(row_data, start=1):
            selected_sheet.cell(row=target_row + i,
                                column=col_idx, value=value)

    # 保存工作簿
    # 替换为您的输出Excel文件路径
    output_file = "/Users/xinwang/Desktop/zhengsiyu/Buried-point-testing/jk埋点.xlsx"
    wb.save(output_file)


result_traspose()
